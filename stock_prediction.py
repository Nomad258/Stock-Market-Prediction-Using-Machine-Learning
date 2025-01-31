import pandas as pd
import yfinance as yf
import numpy as np
import matplotlib.pyplot as plt
import os
import sys
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

try:
    # Ask user for stock ticker and date range
    ticker = input("Enter the stock ticker symbol (e.g., AAPL, TSLA, GOOGL): ").strip().upper()
    start_date = input("Enter the start date (YYYY-MM-DD): ").strip()
    end_date = input("Enter the end date (YYYY-MM-DD): ").strip()

    print("✅ DEBUG: Fetching stock data...")
    stock_data = yf.download(ticker, start=start_date, end=end_date)

    if stock_data.empty:
        print("❌ ERROR: No stock data found. Exiting.")
        exit()

    print("✅ DEBUG: Fetched stock data successfully.")
    print("✅ DEBUG: Creating technical indicators...")

    # Create technical indicators
    stock_data["50_MA"] = stock_data["Close"].rolling(window=50).mean()
    stock_data["200_MA"] = stock_data["Close"].rolling(window=200).mean()
    stock_data["Daily_Return"] = stock_data["Close"].pct_change()

    print("✅ DEBUG: Dropping NaN values and preparing features.")
    # Drop NaN values
    stock_data.dropna(inplace=True)

    # Define features (X) and target (y)
    X = stock_data[["50_MA", "200_MA", "Daily_Return"]]
    y = stock_data["Close"]

    print("✅ DEBUG: Splitting data into training and test sets.")
    # Split data
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, shuffle=False)

    print("✅ DEBUG: Training the Linear Regression model.")
    # Train model
    model = LinearRegression()
    model.fit(X_train, y_train)

    print("✅ DEBUG: Predicting prices.")
    # Predict prices
    y_pred = model.predict(X_test)

    # Ensure y_pred is 1D
    y_pred = y_pred.flatten()

    print("✅ DEBUG: Checking y_pred and y_test shapes before creating DataFrame.")
    y_test = y_test.values.flatten()  # Fix shape issue
    print(f"✅ DEBUG: y_pred shape: {y_pred.shape}")
    print(f"✅ DEBUG: y_test shape: {y_test.shape}")
    print(f"✅ DEBUG: y_pred first 5 values: {y_pred[:5]}")
    print(f"✅ DEBUG: y_test first 5 values: {y_test[:5]}")

    print("✅ DEBUG: Calculating Mean Absolute Error.")
    # Evaluate
    mae = mean_absolute_error(y_test, y_pred)
    print(f"Mean Absolute Error for {ticker}: {mae}")

    # Plot actual vs predicted prices
    plt.figure(figsize=(12,6))
    plt.plot(y_test, label=f"Actual Price ({ticker})")
    plt.plot(y_pred, label=f"Predicted Price ({ticker})", linestyle="dashed")
    plt.legend()
    plt.xlabel("Time")
    plt.ylabel("Stock Price")
    plt.title(f"Stock Price Prediction vs Actual Price ({ticker})")
    graph_path = f"{ticker}_prediction_chart.png"
    plt.savefig(graph_path)
    plt.show(block=False)

    print("✅ DEBUG: Preparing to create DataFrame for Excel.")
    print(f"✅ DEBUG: Expected DataFrame length: {len(y_test)} rows.")

    # Create DataFrame
    predictions_df = pd.DataFrame({
        "Date": X_test.index,  # Use X_test.index instead of y_test.index
        "Actual Price": y_test,
        "Predicted Price": y_pred
    })
    print(f"✅ DEBUG: DataFrame successfully created with shape: {predictions_df.shape}")

    print("✅ DEBUG: DataFrame creation complete. Moving to Excel file creation...")

    print("🔍 DEBUG: About to save Excel file...")
    excel_filename = f"{ticker}_prediction_results.xlsx"
    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        predictions_df.to_excel(writer, sheet_name="Predictions", index=False)

    # Load workbook to format columns and insert graph
    wb = load_workbook(excel_filename)
    ws = wb["Predictions"]

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Insert graph into a new sheet
    ws_chart = wb.create_sheet("Graph")
    img = Image(graph_path)
    ws_chart.add_image(img, "B2")

    # Save final Excel file
    wb.save(excel_filename)

    print(f"✅ Excel file successfully saved with graph: {excel_filename}")

    print("✅ DEBUG: End of script reached.")

except MemoryError:
    print("❌ ERROR: Python ran out of memory while running the script.")
    exit()
except Exception as e:
    print(f"❌ ERROR: An unexpected issue occurred - {str(e)}")
    exit()
